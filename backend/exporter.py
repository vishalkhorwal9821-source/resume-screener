import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import os
import uuid

EXPORT_DIR = "exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

def export_to_excel(session: dict) -> str:
    """Export screening results to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    # Header style
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        "Rank", "Candidate Name", "Final Score (%)", "Bias-Free Score (%)",
        "Recommendation", "Education", "Email", "Phone",
        "Skill Match (%)", "TF-IDF Similarity (%)", "Experience (Years)",
        "Matched Skills", "Missing Skills", "Bias Flags"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for rank, result in enumerate(session["results"], 1):
        row = [
            rank,
            result["candidate_name"],
            result["final_score"],
            result["bias_free_score"],
            result.get("recommendation", ""),
            result.get("education_level", ""),
            result.get("contact_email", ""),
            result.get("contact_phone", ""),
            result["skill_match_score"],
            result["tfidf_similarity"],
            result["experience_years"],
            ", ".join(result["matched_skills"]),
            ", ".join(result["missing_skills"]),
            ", ".join(result["bias_flags"]) if result["bias_flags"] else "None"
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=rank + 1, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True)

        # Color rows by score
        score = result["final_score"]
        if score >= 70:
            fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif score >= 40:
            fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.cell(row=rank + 1, column=col).fill = fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    filepath = os.path.join(EXPORT_DIR, f"screening_{uuid.uuid4().hex[:8]}.xlsx")
    wb.save(filepath)
    return filepath

def export_to_pdf_report(session: dict) -> str:
    """Export screening results to PDF."""
    filepath = os.path.join(EXPORT_DIR, f"report_{uuid.uuid4().hex[:8]}.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph("<b>AI Resume Screening Report</b>", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    jd_preview = session["job_description"][:200] + "..." if len(session["job_description"]) > 200 else session["job_description"]
    jd_para = Paragraph(f"<b>Job Description:</b> {jd_preview}", styles["Normal"])
    elements.append(jd_para)
    elements.append(Spacer(1, 20))

    # Table data
    table_data = [["Rank", "Candidate", "Score", "Recommendation", "Experience", "Matched Skills"]]

    for rank, r in enumerate(session["results"], 1):
        table_data.append([
            str(rank),
            r["candidate_name"],
            f"{r['final_score']}%",
            r.get("recommendation", "Review manually"),
            f"{r['experience_years']} yrs",
            ", ".join(r["matched_skills"][:4]) + ("..." if len(r["matched_skills"]) > 4 else "")
        ])

    table = Table(table_data, colWidths=[30, 120, 50, 70, 70, 160])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F0F9FF"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("WORDWRAP", (0, 0), (-1, -1), True),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))
    footer = Paragraph(f"<i>Generated on {session['created_at'][:10]} | Total Candidates: {len(session['results'])}</i>", styles["Normal"])
    elements.append(footer)

    doc.build(elements)
    return filepath
